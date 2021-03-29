import pandas
hodiny = pandas.read_excel("hodiny.xlsx")

print(hodiny.head())
print(hodiny.shape)

from openpyxl import Workbook
from openpyxl.styles import PatternFill

wb = Workbook()
ws1 = wb.active
ws1.title = "rozvrh"

rozvrh_hodin = [
  ["Anglický jazyk", "Přírodopis", "Dějepis", "Matematika", "Oběd", "Tělesná výchova", "Tělesná výchova", ],
  ["Občanská výchova", "Hudební výchova", "Matematika", "Oběd", "Výtvarná výchova", "Dějepis", ],
  ["Matematika", "Chemie", "Přírodopis", "Fyzika", "Oběd", "Zeměpis", ],
  ["Fyzika", "Anglický jazyk", "Matematika", "Český jazyk", "Dějepis", "Oběd", ],
  ["Český jazyk", "Zeměpis", "Český jazyk", "Výtvarná výchova", "Oběd", ]
]
radek = 1
for den in rozvrh_hodin:
  sloupec = 1
  for predmet in den:
    bunka = ws1.cell(radek, sloupec)
    bunka.value = predmet
    sediva_barva = PatternFill("solid", fgColor="00C0C0C0")
    lahvova_barva = PatternFill("lightDown", fgColor="00008080")
    bila_barva = PatternFill("darkTrellis", fgColor="00FFFFFF")
    if predmet == "Oběd":
      bunka.fill = sediva_barva
    elif predmet == "Výtvarná výchova" or bunka.value == "Hudební výchova":
      bunka.fill = lahvova_barva
    else:
      bunka.fill = bila_barva
    sloupec += 1
  radek += 1

wb.save(filename="rozvrh_hodin.xlsx")

